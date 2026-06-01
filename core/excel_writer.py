"""Template-faithful Excel generation with openpyxl (NEVER pandas).

Loads the clean blank template, clones its single SABLON sheet once per village,
and fills only data cells per the fixed map. Formatting (merges, fonts, wrap,
borders) is preserved by copy_worksheet.
"""
import datetime as _dt
import re
from pathlib import Path
from typing import Optional, Union

from openpyxl import load_workbook
from openpyxl.styles import Alignment

import config
from core.normalize import KoySayfasi

_TEMPLATE_SHEET = "SABLON"
_INVALID_TITLE = re.compile(r"[:\\/?*\[\]]")
_MAX_TITLE = 31


def _safe_title(name: str, used: set[str], suffix: str = "") -> str:
    base = _INVALID_TITLE.sub(" ", name).strip()
    base = re.sub(r"\s+", " ", base) or "Köy"
    room = _MAX_TITLE - len(suffix)
    title = (base[:room].rstrip() + suffix)
    # dedupe
    if title in used:
        n = 2
        while True:
            sfx = f" {n}"
            cand = base[: _MAX_TITLE - len(sfx)].rstrip() + sfx
            if cand not in used:
                title = cand
                break
            n += 1
    used.add(title)
    return title


def _chunkla(kalemler: list, n: int) -> list[list]:
    return [kalemler[i:i + n] for i in range(0, len(kalemler), n)] or [[]]


def _fill_sheet(ws, sayfa_koy: str, kalemler: list, tarih: str, no: Optional[str]):
    ws["E2"] = tarih
    if no:
        ws["G2"] = no if str(no).lower().startswith("no") else f"No: {no}"
    ws["C2"] = config.BIRIM_ADI

    for i, k in enumerate(kalemler):
        r = config.DATA_START_ROW + i
        ws.cell(row=r, column=1).value = i + 1
        ws.cell(row=r, column=2).value = k.kodu or config.KODU_VARSAYILAN
        c = ws.cell(row=r, column=3)
        c.value = f"{sayfa_koy}\n{k.malzeme_adi}"
        a = c.alignment                                   # mandatory line break
        c.alignment = Alignment(horizontal=a.horizontal, vertical=a.vertical, wrap_text=True)
        ws.cell(row=r, column=5).value = k.olcu_birimi or config.OLCU_BIRIMI_VARSAYILAN
        ws.cell(row=r, column=6).value = k.miktar
        # G (karşılanan miktar) intentionally left blank


def yaz(
    sayfalar: list[KoySayfasi],
    tarih: Optional[str] = None,
    no: Optional[str] = None,
    cikti_yolu: Optional[Union[str, Path]] = None,
    template_path: Optional[Union[str, Path]] = None,
) -> Path:
    """Produce ONE new workbook: one sheet per canonical village (overflow ->
    'KöyAdı 2'). Returns the saved path. Every run writes from scratch."""
    if tarih is None:
        tarih = _dt.date.today().strftime("%d.%m.%Y")

    tpath = Path(template_path or config.TEMPLATE_PATH)
    wb = load_workbook(tpath)
    template = wb[_TEMPLATE_SHEET]

    used_titles: set[str] = set()
    if not sayfalar:
        raise ValueError("Yazılacak köy/sayfa yok.")

    for sayfa in sayfalar:
        for ci, chunk in enumerate(_chunkla(sayfa.kalemler, config.MAX_KALEM_PER_SAYFA)):
            suffix = "" if ci == 0 else f" {ci + 1}"
            ws = wb.copy_worksheet(template)
            ws.title = _safe_title(sayfa.koy, used_titles, suffix)
            _fill_sheet(ws, sayfa.koy, chunk, tarih, no)

    # remove the blank template sheet, keep only filled village sheets
    del wb[_TEMPLATE_SHEET]

    if cikti_yolu is None:
        ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        cikti_yolu = config.OUTPUT_DIR / f"tasinir_{ts}.xlsx"
    cikti_yolu = Path(cikti_yolu)
    wb.save(cikti_yolu)
    return cikti_yolu
