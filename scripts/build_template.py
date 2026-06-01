"""One-time setup.

Derive a CLEAN single-page blank template (templates/bos_sablon.xlsx) from the
source workbook, and extract the canonical village list (data/koy_listesi.csv)
from its sheet names.

The source data area is hand-built and irregular (mixed 1- and 2-row merged
items, sporadic SIRA NO merges). We keep the header band (rows 1-6) and the
column widths exactly, then rebuild a REGULAR one-row-per-item grid starting at
DATA_START_ROW, followed by a config-driven footer (Mevlüt DUMAN; Kıvanç AKTAN
removed).

Run:  python scripts/build_template.py
"""
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config  # noqa: E402

BASE_SHEET = "gaziköy"   # representative sheet to inherit the header band from
N_DATA_ROWS = config.MAX_KALEM_PER_SAYFA  # 25 regular item rows
START = config.DATA_START_ROW             # 7

MED = Side(style="medium")
BOX = Border(top=MED, bottom=MED, left=MED, right=MED)


def _box_edges(ws, row):
    """Keep the outer table box: medium on column A's left and G's right."""
    a = ws.cell(row=row, column=1)
    g = ws.cell(row=row, column=7)
    a.border = Border(left=MED, top=a.border.top, bottom=a.border.bottom, right=a.border.right)
    g.border = Border(right=MED, top=g.border.top, bottom=g.border.bottom, left=g.border.left)


def _set(ws, coord, value, font, align, *, merge=None, border=None):
    c = ws[coord]
    c.value = value
    c.font = font
    c.alignment = align
    if border is not None:
        c.border = border
    if merge:
        ws.merge_cells(merge)
    return c


def build_data_grid(ws):
    f_sira = Font(name="Arial", size=10)
    f_kod = Font(name="Arial", size=10)
    f_adi = Font(name="Arial", size=8, bold=True)
    f_birim = Font(name="Arial", size=9, bold=True)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i in range(N_DATA_ROWS):
        r = START + i
        ws.row_dimensions[r].height = 28.5
        # A SIRA NO, B KODU
        _set(ws, f"A{r}", None, f_sira, center_wrap, border=BOX)
        _set(ws, f"B{r}", None, f_kod, center_wrap, border=BOX)
        # C:D merged ADI (koy\nmalzeme), wrap mandatory
        ws.merge_cells(f"C{r}:D{r}")
        _set(ws, f"C{r}", None, f_adi, center_wrap, border=BOX)
        ws[f"D{r}"].border = BOX
        # E birim, F istenilen miktar, G karşılanan (blank)
        _set(ws, f"E{r}", None, f_birim, center_wrap, border=BOX)
        _set(ws, f"F{r}", None, f_birim, center_wrap, border=BOX)
        _set(ws, f"G{r}", None, f_sira, center_wrap, border=BOX)


def build_footer(ws):
    fc = config.FOOTER
    m = config.FOOTER_METIN
    f9 = Font(name="Arial", size=9)
    f9b = Font(name="Arial", size=9, bold=True)
    f10b = Font(name="Arial", size=10, bold=True)
    cc = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lc = Alignment(horizontal="left", vertical="center", wrap_text=True)

    r0 = START + N_DATA_ROWS  # first footer row, right below the grid

    # explanatory band (2 columns), still inside the box
    ws.row_dimensions[r0].height = 30
    _set(ws, f"A{r0}", m["istek_sol"], f9, cc, merge=f"A{r0}:C{r0}")
    _set(ws, f"D{r0}", m["teslim_sag"], f9, cc, merge=f"D{r0}:G{r0}")

    # signature block (left = istek yapan, right = kayıt yetkilisi)
    rows = [
        (m["label_istek"], m["label_kayit"], f10b, cc),
        (f"Adı, Soyadı : {fc['istek_yapan']['ad']}",
         f"Adı, Soyadı : {fc['kayit_yetkilisi']['ad']}", f10b, lc),
        (f"Unvanı        :{fc['istek_yapan']['unvan']}",
         f"Unvanı        : {fc['kayit_yetkilisi']['unvan']}", f10b, lc),
        (m["imza"], m["imza"], f10b, lc),
        (None, None, f10b, lc),  # spacer inside box
    ]
    for j, (lval, rval, fnt, aln) in enumerate(rows):
        r = r0 + 1 + j
        _set(ws, f"A{r}", lval, fnt, aln, merge=f"A{r}:C{r}")
        _set(ws, f"D{r}", rval, fnt, aln, merge=f"D{r}:G{r}")

    # birim yöneticisi block (centered, B:D)
    ry = r0 + 6
    ybloc = [
        (m["label_yonetici"], cc),
        (f"Adı, Soyadı : {fc['birim_yoneticisi']['ad']}", lc),
        (f"Unvanı        : {fc['birim_yoneticisi']['unvan']}", lc),
        (m["imza"], lc),
    ]
    for j, (val, aln) in enumerate(ybloc):
        r = ry + j
        _set(ws, f"B{r}", val, f10b, aln, merge=f"B{r}:D{r}")

    # closing line + footnote
    r_close = ry + 4
    _set(ws, f"A{r_close}", m["belge_ornegi"], f9,
         Alignment(horizontal="left", vertical="center", wrap_text=True),
         merge=f"A{r_close}:G{r_close}")
    ws.row_dimensions[r_close].height = 24

    r_tmy = r_close + 2
    _set(ws, f"A{r_tmy}", m["tmy"], f9b,
         Alignment(horizontal="left", vertical="center"))

    # keep the outer table box continuous from grid top through closing line,
    # and close the bottom on the closing line.
    for r in range(r0, r_close + 1):
        _box_edges(ws, r)
    for col in range(1, 8):
        c = ws.cell(row=r_close, column=col)
        c.border = Border(left=c.border.left, right=c.border.right,
                          top=c.border.top, bottom=MED)
    # re-assert L/R medium on the closing row corners
    _box_edges(ws, r_close)
    ws.cell(row=r_close, column=1).border = Border(left=MED, right=ws.cell(row=r_close, column=1).border.right, top=ws.cell(row=r_close, column=1).border.top, bottom=MED)
    ws.cell(row=r_close, column=7).border = Border(right=MED, left=ws.cell(row=r_close, column=7).border.left, top=ws.cell(row=r_close, column=7).border.top, bottom=MED)


def build_template():
    wb = load_workbook(config.SOURCE_XLSX)
    base = wb[BASE_SHEET]

    # drop every other sheet, keep only the representative one
    for name in list(wb.sheetnames):
        if name != BASE_SHEET:
            del wb[name]

    ws = base
    ws.title = "SABLON"

    # unmerge everything from the data area down, then delete those rows so we
    # can rebuild a clean grid. Header band (rows 1-6) is left untouched.
    for mr in [str(m) for m in list(ws.merged_cells.ranges) if m.min_row >= START]:
        ws.unmerge_cells(mr)
    if ws.max_row >= START:
        ws.delete_rows(START, ws.max_row - START + 1)

    # clear stale per-sheet header fields (date, no) -> filled at write time
    ws["E2"] = None
    ws["G2"] = config.NO_VARSAYILAN
    ws["C2"] = config.BIRIM_ADI

    build_data_grid(ws)
    build_footer(ws)

    config.TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(config.TEMPLATE_PATH)
    print(f"[OK] Temiz şablon yazıldı: {config.TEMPLATE_PATH}")
    print(f"     Veri satırları: {START}..{START + N_DATA_ROWS - 1}  ({N_DATA_ROWS} kalem)")


def build_village_list():
    wb = load_workbook(config.SOURCE_XLSX, read_only=True)
    names = sorted({n.strip() for n in wb.sheetnames if n.strip()},
                   key=lambda s: s.lower())
    wb.close()
    config.KOY_LISTESI_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(config.KOY_LISTESI_PATH, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["koy"])
        for n in names:
            w.writerow([n])
    print(f"[OK] Köy listesi yazıldı: {config.KOY_LISTESI_PATH}  ({len(names)} köy)")


if __name__ == "__main__":
    if not config.SOURCE_XLSX.exists():
        raise SystemExit(f"Kaynak dosya bulunamadı: {config.SOURCE_XLSX}")
    build_template()
    build_village_list()
