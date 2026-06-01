"""Golden-file tests for the openpyxl writer: cell map, wrap, merges, footer,
sheet titles, and the 25-row overflow rule."""
import datetime as _dt

import pytest
from openpyxl import load_workbook

import config
from core import excel_writer
from core.normalize import KoySayfasi
from core.schema import Kalem


def _k(koy, malz, mik, birim="Ad", kodu=None):
    return Kalem(koy=koy, malzeme_adi=malz, miktar=mik, olcu_birimi=birim, kodu=kodu,
                 guven_koy=1.0, guven_malzeme=1.0, guven_miktar=1.0)


@pytest.fixture(scope="module")
def template_var():
    if not config.TEMPLATE_PATH.exists():
        pytest.skip("templates/bos_sablon.xlsx yok — önce scripts/build_template.py çalıştır.")


def test_hucre_haritasi_ve_wrap(tmp_path, template_var):
    sayfa = KoySayfasi(koy="Merkez Gaziköy", kalemler=[
        _k("Merkez Gaziköy", "Solar Şarj Ünitesi", 1),
        _k("Merkez Gaziköy", "Dalgıç Pompa", 2, birim="Ad"),
    ])
    out = tmp_path / "g.xlsx"
    excel_writer.yaz([sayfa], tarih="01.06.2026", no="12", cikti_yolu=out)

    wb = load_workbook(out)
    assert "SABLON" not in wb.sheetnames           # template removed
    ws = wb["Merkez Gaziköy"]
    r = config.DATA_START_ROW
    assert ws.cell(row=r, column=1).value == 1     # SIRA NO
    assert ws.cell(row=r, column=2).value == config.KODU_VARSAYILAN
    c = ws.cell(row=r, column=3)
    assert c.value == "Merkez Gaziköy\nSolar Şarj Ünitesi"
    assert c.alignment.wrap_text is True            # mandatory line break
    assert ws.cell(row=r, column=5).value == "Ad"
    assert ws.cell(row=r, column=6).value == 1
    assert ws.cell(row=r, column=7).value is None   # KARŞILANAN boş
    # second item, renumbered
    assert ws.cell(row=r + 1, column=1).value == 2
    # header date + no
    assert ws["E2"].value == "01.06.2026"
    assert "12" in str(ws["G2"].value)


def test_cd_merge_korunur(tmp_path, template_var):
    sayfa = KoySayfasi(koy="Köy A", kalemler=[_k("Köy A", "Vana", 1)])
    out = tmp_path / "m.xlsx"
    excel_writer.yaz([sayfa], cikti_yolu=out)
    ws = load_workbook(out)["Köy A"]
    r = config.DATA_START_ROW
    merged = {str(m) for m in ws.merged_cells.ranges}
    assert f"C{r}:D{r}" in merged


def test_footer_mevlut_var_kivanc_yok(tmp_path, template_var):
    sayfa = KoySayfasi(koy="Köy B", kalemler=[_k("Köy B", "X", 1)])
    out = tmp_path / "f.xlsx"
    excel_writer.yaz([sayfa], cikti_yolu=out)
    ws = load_workbook(out)["Köy B"]
    metin = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "Mevlüt DUMAN" in metin
    assert "Kıvanç" not in metin
    assert "Burak ARSLAN" in metin
    assert "İsmail TAŞTAN" in metin


def test_25_satir_tasmasi_ikinci_sayfa(tmp_path, template_var):
    kalemler = [_k("Büyükköy", f"Malzeme {i}", 1) for i in range(26)]
    sayfa = KoySayfasi(koy="Büyükköy", kalemler=kalemler)
    out = tmp_path / "t.xlsx"
    excel_writer.yaz([sayfa], cikti_yolu=out)
    wb = load_workbook(out)
    assert "Büyükköy" in wb.sheetnames
    assert "Büyükköy 2" in wb.sheetnames
    # first sheet has 25, second has 1
    s1, s2 = wb["Büyükköy"], wb["Büyükköy 2"]
    assert s1.cell(row=config.DATA_START_ROW + 24, column=1).value == 25
    assert s2.cell(row=config.DATA_START_ROW, column=1).value == 1


def test_sayfa_adi_temizleme(tmp_path, template_var):
    # invalid chars + over-long name
    uzun = "Merkez/Köy: Çok Uzun Bir Köy Adı [test] Daha Da Uzun"
    sayfa = KoySayfasi(koy=uzun, kalemler=[_k(uzun, "X", 1)])
    out = tmp_path / "s.xlsx"
    excel_writer.yaz([sayfa], cikti_yolu=out)
    wb = load_workbook(out)
    title = wb.sheetnames[0]
    assert len(title) <= 31
    assert not any(ch in title for ch in ':\\/?*[]')


def test_cogul_koy_cogul_sekme(tmp_path, template_var):
    sayfalar = [
        KoySayfasi(koy="Köy X", kalemler=[_k("Köy X", "A", 1)]),
        KoySayfasi(koy="Köy Y", kalemler=[_k("Köy Y", "B", 1)]),
    ]
    out = tmp_path / "c.xlsx"
    excel_writer.yaz(sayfalar, cikti_yolu=out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Köy X", "Köy Y"}
